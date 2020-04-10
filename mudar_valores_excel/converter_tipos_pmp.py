# Script para modificar valores em planilhas do Excel
# -> O script roda em todas as planilhas num diretório
# USE COM CUIDADO - de preferência não nos arquivos originais (use cópias)

import os
import openpyxl
from openpyxl.styles import Alignment
import re


# Parâmetros:
# path: diretório dos arquivos (UNICODE: use / para separar os dirs)
# lista_ext: formatos que serão abertos
# lin_proc_v: valor que identifica a linha procurada
# col_proc_v: valor(es) que serão substituídos
#novo_valor: valor que irá substituir
#base_ref: base de referência
path = 'C:/Users/camilo.sidou/Desktop/PMPS2/mod' 
lista_ext = ['.xlsx']

lista_path = []

for root, dir, file in os.walk(path):
    for i in file:
        for ext in lista_ext:
            if ext in i:
                arq_path = [i, os.path.join(root, i)]
                lista_path.append(arq_path)

idx = 0
for i in lista_path:
    try:
        arq = openpyxl.load_workbook(filename=i[1])
        plan = arq.active
        max_lin = plan.max_row
        max_col = plan.max_column
    except:
        print('>> ERRO (não conseguiu abrir):', i[1])
        continue
    else:
        for lin in range(1,max_lin+1):
            for col in range(1,max_col):
                if (lin == 9 and col >= 3) or (lin >= 13 and col >= 3):
                    plan[lin][col].alignment = Alignment(wrapText=False)
                    plan[lin][col].number_format = '0.00'
        print(i[0], "concluído!")
        arq.save(i[1])