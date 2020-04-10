import openpyxl as pl
import os
import re
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time

arq_in = 'C:/Users/camilo.sidou/Google Drive/DICAR/scripts/sigtap_web_crawler/saida_2.xlsx'
arq_out = 'C:/Users/camilo.sidou/Google Drive/DICAR/scripts/sigtap_web_crawler/saida_2.xlsx'

arq = pl.load_workbook(arq_in)
plan = arq.active

driver = webdriver.Chrome('C:/Users/camilo.sidou/Google Drive/DICAR/chromedriver.exe')
current_path = os.getcwd()
driver_opt = webdriver.ChromeOptions()
driver_opt.add_argument('download.default_directory=' + current_path)

n_lin = plan.max_row
for lin in range(1, n_lin):
    if (plan.cell(row = lin, column = 3).value is not None):
        continue    
    sigtap = str(plan.cell(row = lin, column = 1).value)
    print("> SIGTAP:", sigtap)
    driver.get('http://sigtap.datasus.gov.br/tabela-unificada/app/sec/inicio.jsp')
    time.sleep(2)
    acessar_tabela = None
    try:
        acessar_tabela = driver.find_element_by_id('acessoAutomatico')
        if (acessar_tabela is not None):
            time.sleep(2)
            acessar_tabela.click()
    except:
        pass
    form_cod = None
    time.sleep(2)
    try:
        form_cod = driver.find_element_by_id('formConsultarProcedimento:codigo')
    except:
        continue
    form_cod.send_keys(Keys.CONTROL + 'a')
    form_cod.send_keys(Keys.DELETE)
    arq.save(arq_out)
    form_cod.send_keys(sigtap)
    form_enviar = None
    try:
        form_enviar = driver.find_element_by_id('formConsultarProcedimento:localizar').click()
    except:
        continue
    try:
        nome_proc = driver.find_element_by_id('formConsultarProcedimento:historicoProcedimento:0:_idJsp258').text
    except:
        continue
    nome_proc = nome_proc[17:]
    time.sleep(2)
    plan.cell(row = lin, column = 3).value = str(nome_proc)
    print(">> Escrito:", nome_proc)
    time.sleep(5)
driver.close()
arq.save(arq_out)