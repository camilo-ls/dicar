import openpyxl as pl
import PyPDF2 as pf
import re

arq = 'C:/Users/camilo.sidou/Google Drive/DICAR/scripts/sigtap.pdf'
arq_saida = 'C:/Users/camilo.sidou/Google Drive/DICAR/scripts/saida.xlsx'
re_sigtap = re.compile(r'0\d\d\d\d\d\d\d\d\d')
re_nome_proc = re.compile(r'->\s+.*\n.*')
# 0101010010 -> ATIVIDADE EDUCATIVA/ORIENTACAO EM GRUPO NA ATENCAO BASICA

arq_pdf = open(arq, 'rb')
reader = pf.PdfFileReader(arq_pdf)
encontrados_sigtap = []
encontrados_nome_proc = []

for i in range(reader.numPages):
    res_f = []
    page = reader.getPage(i)
    page_text = page.extractText()
    #print(page_text)
    res_sigtap = re_sigtap.findall(page_text)
    res_nome_proc = re_nome_proc.findall(page_text)
    for j in res_sigtap:
        encontrados_sigtap.append(j)
        print('Encontrado:', j)
    for j in res_nome_proc:
        j = j.replace('\n', '').replace('\r', '').replace('-', '')
        j = j[2:]
        encontrados_nome_proc.append(j)
        print('Encontrado:', j)

arq_pdf.close()
arq_excel = pl.Workbook()
plan = arq_excel.active

idx_row = 1
for i in encontrados_sigtap:
    plan.cell(row=idx_row, column=1).value = i
    print('Escrito:', i)
    idx_row += 1

idx_row = 1
for i in encontrados_nome_proc:
    plan.cell(row=idx_row, column=2).value = i
    print('Escrito:', i)
    idx_row += 1

arq_excel.save(arq_saida)


